"""
Excel数据导出功能
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from app.models import Game, GamePlayer, GameRound, User
from app.game_logic import FINAL_ENV_POSITIVE_RATE, FINAL_ENV_NEGATIVE_RATE
from typing import List
import os


def _excel_column_letter(col_index: int) -> str:
    """1-based column index to Excel column letter: 1->A, 26->Z, 27->AA, ..."""
    result = ""
    while col_index > 0:
        col_index, r = (col_index - 1) // 26, (col_index - 1) % 26
        result = chr(65 + r) + result
    return result or "A"


def write_game_to_sheet(ws, db: Session, game_id: int):
    """
    将单局游戏数据写入已有工作表。含：每轮 NT/ENV/选择，6–10 轮申领补贴，11–15 轮申领补贴+投票（谁都不选记 0）。
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.models import GameVote
    game = db.query(Game).filter(Game.id == game_id).first()
    if not game:
        raise ValueError(f"游戏 {game_id} 不存在")
    players = db.query(GamePlayer).filter(GamePlayer.game_id == game_id).all()
    if not players:
        raise ValueError("游戏没有玩家")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    headers = ["玩家", "用户名"]
    for round_num in range(1, 16):
        headers.append(f"Round{round_num} NT")
        headers.append(f"Round{round_num} ENV")
        headers.append(f"Round{round_num} 选择")
        if round_num >= 6:
            headers.append(f"Round{round_num} 申领补贴")
        if round_num >= 11:
            headers.append(f"Round{round_num} 投票")
    headers.extend(["NT(结算前)", "最终ENV", "生态结算", "最终NT", "总收益", "是否获胜"])
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    all_rounds = db.query(GameRound).filter(GameRound.game_id == game_id).order_by(
        GameRound.round_number, GameRound.player_id
    ).all()
    player_rounds = {}
    for round_data in all_rounds:
        if round_data.player_id not in player_rounds:
            player_rounds[round_data.player_id] = {}
        player_rounds[round_data.player_id][round_data.round_number] = round_data
    votes_list = db.query(GameVote).filter(GameVote.game_id == game_id).all()
    votes_map = {}
    for v in votes_list:
        votes_map[(v.round_number, v.voter_id)] = v.target_id if v.target_id is not None else 0
    for row_idx, player in enumerate(players, 2):
        if getattr(player, "username", None) and str(player.username).strip():
            username = (player.username or "").strip()
        elif getattr(player, "user_id", None):
            user = db.query(User).filter(User.id == player.user_id).first()
            username = user.username if user else f"玩家{player.id}"
        else:
            username = f"玩家{player.id}"
        ws.cell(row=row_idx, column=1, value=player.id)
        ws.cell(row=row_idx, column=2, value=username)
        col_idx = 3
        for round_num in range(1, 16):
            if player.id in player_rounds and round_num in player_rounds[player.id]:
                rd = player_rounds[player.id][round_num]
                ws.cell(row=row_idx, column=col_idx, value=round(rd.nt_after, 1))
                ws.cell(row=row_idx, column=col_idx + 1, value=round(rd.env_after, 1))
                ws.cell(row=row_idx, column=col_idx + 2, value="有机" if rd.choice == "organic" else "无机")
                col_idx += 3
                if round_num >= 6:
                    ws.cell(row=row_idx, column=col_idx, value="是" if rd.applied_subsidy else "否")
                    col_idx += 1
                if round_num >= 11:
                    vote_val = votes_map.get((round_num, player.id), "")
                    ws.cell(row=row_idx, column=col_idx, value=vote_val if vote_val != "" else 0)
                    col_idx += 1
            else:
                ws.cell(row=row_idx, column=col_idx, value="-")
                ws.cell(row=row_idx, column=col_idx + 1, value="-")
                ws.cell(row=row_idx, column=col_idx + 2, value="-")
                col_idx += 3
                if round_num >= 6:
                    ws.cell(row=row_idx, column=col_idx, value="-")
                    col_idx += 1
                if round_num >= 11:
                    ws.cell(row=row_idx, column=col_idx, value="-")
                    col_idx += 1
        final_nt = player.final_nt if player.final_nt else player.current_nt
        final_env = player.final_env if player.final_env is not None else player.current_env
        env_settlement = (final_env * FINAL_ENV_POSITIVE_RATE if final_env > 0 else final_env * FINAL_ENV_NEGATIVE_RATE) if final_env is not None else 0
        nt_before_settlement = final_nt - env_settlement if final_nt is not None else player.current_nt
        ws.cell(row=row_idx, column=col_idx, value=round(nt_before_settlement, 1))
        ws.cell(row=row_idx, column=col_idx + 1, value=round(final_env, 1) if final_env is not None else "-")
        ws.cell(row=row_idx, column=col_idx + 2, value=round(env_settlement, 1))
        ws.cell(row=row_idx, column=col_idx + 3, value=round(final_nt, 1) if final_nt is not None else "-")
        total_reward = (final_nt or player.current_nt) - player.initial_nt
        ws.cell(row=row_idx, column=col_idx + 4, value=round(total_reward, 1))
        ws.cell(row=row_idx, column=col_idx + 5, value="是" if getattr(player, "is_winner", False) else "否")
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[_excel_column_letter(col)].width = 12


def export_batch_to_excel(db: Session, game_ids: List[int], output_path: str):
    """将多局游戏导出到同一 Excel 文件，每局一页（sheet）。"""
    wb = Workbook()
    wb.remove(wb.active)
    for idx, game_id in enumerate(game_ids, 1):
        ws = wb.create_sheet(title=f"游戏{idx}", index=idx - 1)
        write_game_to_sheet(ws, db, game_id)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def export_game_to_excel(db: Session, game_id: int, output_path: str = None):
    """
    导出单局游戏数据到Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "游戏数据"
    write_game_to_sheet(ws, db, game_id)
    if output_path is None:
        output_dir = "exports"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"game_{game_id}_data.xlsx")
    wb.save(output_path)
    return output_path
